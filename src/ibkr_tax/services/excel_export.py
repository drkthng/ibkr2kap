from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload
from ibkr_tax.schemas.report import TaxReport, CombinedTaxReport
from ibkr_tax.models.database import Account, Gain, Trade, FXGain, FXFIFOLot, CashTransaction

EXCEL_STRINGS = {
    "en": {
        "summary_title": "Anlage KAP Summary",
        "kap_report_title": "IBKR2KAP — Anlage KAP Report",
        "accounts_label": "Accounts: ",
        "account_label": "Account: ",
        "tax_year_label": "Tax Year: ",
        "col_line": "Line",
        "col_desc": "Description",
        "col_amount": "Amount (EUR)",
        "sheet_matched_stocks": "Stock Sales (Mat.)",
        "sheet_matched_options": "Derivatives (Mat.)",
        "sheet_dividends": "Dividends, Interest & Other",
        "sheet_margin": "Margin Costs (Info)",
        "sheet_deposits": "Deposits & Withdrawals (Info)",
        "sheet_fx": "FX Gains (§ 23 EStG)",
        "sheet_audit": "Transaction List (All)",
        "margin_warning": "⚠️ Margin interest (Broker Interest Paid) is NOT deductible according to § 20 para. 9 EStG and is NOT included in Anlage KAP.",
        "margin_total_label": "Total Margin Interest (EUR):",
        "header_settle_date": "Settle Date",
        "header_acq_date": "Acquisition Date",
        "header_symbol": "Symbol",
        "header_qty": "Quantity",
        "header_proceeds_brutto": "Proceeds (Gross EUR)",
        "header_cost_brutto": "Cost (Gross EUR)",
        "header_comm": "Commission (EUR)",
        "header_gain_loss": "Gain/Loss (EUR)",
        "header_account": "Account",
        "header_description": "Description",
        "header_type": "Type",
        "header_currency": "Currency",
        "header_amount_brutto": "Amount (Gross)",
        "header_fx_rate": "FX Rate",
        "header_amount_eur": "Amount (EUR)",
        "header_wht_eur": "Withholding Tax (EUR)",
        "header_date": "Date",
        "header_amount": "Amount",
        "header_disposal_date": "Disposal Date",
        "header_days_held": "Days Held",
        "header_proceeds_eur": "Proceeds (EUR)",
        "header_cost_eur": "Cost (EUR)",
        "header_tax_relevant": "Tax Relevant?",
        "header_trade_id": "Trade ID",
        "header_cat": "Cat",
        "header_buy_sell": "Buy/Sell",
        "header_open_close": "Open/Close",
        "header_price": "Price",
        "header_taxes_eur": "Taxes (EUR)",
        "yes": "YES",
        "no": "NO",
        "breakdown_account": "BREAKDOWN: Account {}",
    },
    "de": {
        "summary_title": "Anlage KAP Übersicht",
        "kap_report_title": "IBKR2KAP — Anlage KAP Bericht",
        "accounts_label": "Konten: ",
        "account_label": "Konto: ",
        "tax_year_label": "Steuerjahr: ",
        "col_line": "Zeile",
        "col_desc": "Bezeichnung",
        "col_amount": "Betrag (EUR)",
        "sheet_matched_stocks": "Aktienveräußerungen (Mat.)",
        "sheet_matched_options": "Termingeschäfte (Mat.)",
        "sheet_dividends": "Dividenden, Zinsen & Sonstiges",
        "sheet_margin": "Marginkosten (Info)",
        "sheet_deposits": "Ein- und Auszahlungen (Info)",
        "sheet_fx": "Währungsgewinne (§ 23 EStG)",
        "sheet_audit": "Transaktionsliste (Alle)",
        "margin_warning": "⚠️ Marginzinsen (Broker Interest Paid) sind gemäß § 20 Abs. 9 EStG nicht als Werbungskosten abzugsfähig und fließen NICHT in die Anlage KAP ein.",
        "margin_total_label": "Gesamt Marginzinsen (EUR):",
        "header_settle_date": "Valuta-Datum",
        "header_acq_date": "Anschaffungsdatum",
        "header_symbol": "Symbol",
        "header_qty": "Menge",
        "header_proceeds_brutto": "Erlös (Brutto EUR)",
        "header_cost_brutto": "Kosten (Brutto EUR)",
        "header_comm": "Spesen (EUR)",
        "header_gain_loss": "Gewinn/Verlust (EUR)",
        "header_account": "Konto",
        "header_description": "Beschreibung",
        "header_type": "Typ",
        "header_currency": "Währung",
        "header_amount_brutto": "Betrag (Brutto)",
        "header_fx_rate": "FX Rate",
        "header_amount_eur": "Betrag (EUR)",
        "header_wht_eur": "Quellensteuer (EUR)",
        "header_date": "Datum",
        "header_amount": "Betrag",
        "header_disposal_date": "Datum Dispo",
        "header_days_held": "Haltedauer (Tage)",
        "header_proceeds_eur": "Erlös (EUR)",
        "header_cost_eur": "Kosten (EUR)",
        "header_tax_relevant": "Steuerrelevant?",
        "header_trade_id": "Trade ID",
        "header_cat": "Kat",
        "header_buy_sell": "Kauf/Verkauf",
        "header_open_close": "Open/Close",
        "header_price": "Preis",
        "header_taxes_eur": "Steuern (EUR)",
        "yes": "JA",
        "no": "NEIN",
        "breakdown_account": "BREAKDOWN: Konto {}",
    }
}


class ExcelExportService:
    def __init__(self, session: Session):
        self.session = session

        # Formatting placeholders (initialized in export())
        self.bold_font: Font | None = None
        self.title_font: Font | None = None
        self.header_fill: PatternFill | None = None
        self.euro_format: str = ""
        self.qty_format: str = ""
        self.date_format: str = ""
        self.TR = EXCEL_STRINGS["de"] # Default



    def _init_formatting(self):
        """Initializes Excel formatting constants."""
        self.bold_font = Font(bold=True)
        self.title_font = Font(bold=True, size=14)
        self.header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        self.euro_format = '#,##0.00 €'
        self.qty_format = '#,##0.0000'
        self.date_format = 'yyyy-mm-dd'

    def export(self, report: TaxReport, output_path: str, lang: str = "de") -> None:
        """
        Produces an elegantly formatted Excel report with full row-level transparency.
        """
        self.TR = EXCEL_STRINGS.get(lang, EXCEL_STRINGS["de"])
        self._init_formatting()
        wb = Workbook()
        
        # 1-8. Normal Sheets
        self._add_summary_sheet(wb, report)
        self._add_matched_gains_sheet(wb, report, self.TR["sheet_matched_stocks"], "Aktien")
        self._add_matched_gains_sheet(wb, report, self.TR["sheet_matched_options"], "Termingeschäfte")
        self._add_cash_details_sheet(wb, report)
        self._add_margin_interest_sheet(wb, report)
        self._add_deposits_withdrawals_sheet(wb, report)
        self._add_fx_gains_sheet(wb, report)
        self._add_audit_trail_sheet(wb, report)

        wb.save(output_path)

    def export_combined(self, combined_report: CombinedTaxReport, output_path: str, lang: str = "de") -> None:
        """
        Produces a combined Excel report for multiple accounts.
        """
        self.TR = EXCEL_STRINGS.get(lang, EXCEL_STRINGS["de"])
        self._init_formatting()
        wb = Workbook()

        # 1. Summary Sheet (Combined + Individual)
        self._add_summary_sheet(wb, combined_report)

        # 2-8. Detail Sheets with "Konto" column enabled
        self._add_matched_gains_sheet(wb, combined_report, self.TR["sheet_matched_stocks"], "Aktien", show_account=True)
        self._add_matched_gains_sheet(wb, combined_report, self.TR["sheet_matched_options"], "Termingeschäfte", show_account=True)
        self._add_cash_details_sheet(wb, combined_report, show_account=True)
        self._add_margin_interest_sheet(wb, combined_report, show_account=True)
        self._add_deposits_withdrawals_sheet(wb, combined_report, show_account=True)
        self._add_fx_gains_sheet(wb, combined_report, show_account=True)
        self._add_audit_trail_sheet(wb, combined_report, show_account=True)

        wb.save(output_path)

    def _add_summary_sheet(self, wb, report):
        ws = wb.active
        ws.title = self.TR["summary_title"]
        
        is_combined = isinstance(report, CombinedTaxReport)
        
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_suffix = f" ({self.TR['yes']})" if is_combined else "" # Simplified
        title_cell.value = self.TR["kap_report_title"] + title_suffix
        title_cell.font = self.title_font
        title_cell.alignment = Alignment(horizontal="center")
        
        if is_combined:
            ws["A2"] = self.TR["accounts_label"] + ", ".join(report.account_ids)
        else:
            ws["A2"] = self.TR["account_label"] + report.account_id
        ws["B2"] = self.TR["tax_year_label"] + str(report.tax_year)
        
        headers = [self.TR["col_line"], self.TR["col_desc"], self.TR["col_amount"]]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill
            
        def _write_kap_rows(start_row, r_data):
            curr_row = start_row
            for zeile, desc, val in r_data:
                ws.cell(row=curr_row, column=1).value = zeile
                ws.cell(row=curr_row, column=2).value = desc
                val_cell = ws.cell(row=curr_row, column=3)
                val_cell.value = val
                if isinstance(val, (Decimal, float)):
                    val_cell.number_format = self.euro_format
                curr_row += 1
            return curr_row

        def _get_report_rows(rep):
            # These descriptions are mostly German tax terms, keeping them standard
            # but using English where appropriate if needed. For now using the existing strings
            # and potentially translating them if I want to be 100% thorough.
            # But the user specifically wanted to translate MISSING parts to German.
            # Since the Excel was already mostly German, I'll keep the German descriptions here
            # but use self.TR for the structural parts.
            return [
                ("7", "Kapitalerträge (Dividenden / Zinsen / Ausgleichszahlungen / Sonstige)", rep.kap_line_7_kapitalertraege),
                ("8", "Aktien-Veräußerungsgewinne", rep.kap_line_8_gewinne_aktien),
                ("9", "Aktien-Veräußerungsverluste", rep.kap_line_9_verluste_aktien),
                ("10", "Termingeschäfte (netto)", rep.kap_line_10_termingeschaefte),
                ("", "  - davon Gewinne", rep.kap_termingeschaefte_gains),
                ("", "  - davon Verluste", rep.kap_termingeschaefte_losses),
                ("15", "Anrechenbare ausländische Steuern", rep.kap_line_15_quellensteuer),
                ("", "", ""),
                ("SO", "Fremdwährungsgeschäfte (§ 23 EStG)", ""),
                ("", "  - Gesamtgewinn/-verlust", rep.so_fx_gains_total),
                ("", "  - Davon steuerpflichtig (< 1 Jahr)", rep.so_fx_gains_taxable_1y),
                ("", "  - Davon steuerfrei (> 1 Jahr)", rep.so_fx_gains_tax_free),
                ("", f"  - Freigrenze (1000€) unterschritten?", self.TR["yes"] if rep.so_fx_freigrenze_applies else self.TR["no"]),
                ("", "", ""),
                ("", "Zusammenfassung nach Verlusttöpfen (§ 20 Abs. 6 EStG)", ""),
                ("", "Aktientopf (Netto: Zeile 8 - 9)", rep.aktien_net_result),
                ("", "  Hinweis: Nur mit Aktiengewinnen verrechenbar.", ""),
                ("", "Allgemeiner Topf (Zeile 7 + 10)", rep.allgemeiner_topf_result),
                ("", "  - davon Dividenden & Zinsen", rep.dividends_interest_total),
                ("", "  - davon Sonstige Kursgewinne (ETFs, etc.)", rep.sonstige_gains_total),
                ("", "  - davon Termingeschäfte (Gains)", rep.kap_termingeschaefte_gains),
                ("", "  - davon Termingeschäfte (Losses)", rep.kap_termingeschaefte_losses),
                ("", "  - davon Termingeschäfte (Netto)", rep.kap_line_10_termingeschaefte),
                ("", "", ""),
                ("", "Marginkosten (Info, nicht abzugsfähig)", rep.margin_interest_paid),
            ]

        # 1. Main Summary (Combined or Single)
        next_r = _write_kap_rows(5, _get_report_rows(report))

        # 2. Individual Summaries (if combined)
        if is_combined:
            for acc_rep in report.per_account_reports:
                next_r += 1
                ws.cell(row=next_r, column=2).value = self.TR["breakdown_account"].format(acc_rep.account_id)
                ws.cell(row=next_r, column=2).font = self.bold_font
                next_r += 1
                next_r = _write_kap_rows(next_r, _get_report_rows(acc_rep))
            
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 20

    def _add_matched_gains_sheet(self, wb, report, title, pool_filter, show_account=False):
        ws = wb.create_sheet(title)
        headers = [
            self.TR["header_settle_date"], self.TR["header_acq_date"], self.TR["header_symbol"], self.TR["header_qty"], 
            self.TR["header_proceeds_brutto"], self.TR["header_cost_brutto"], self.TR["header_comm"], self.TR["header_gain_loss"]
        ]
        if show_account:
            headers = [self.TR["header_account"]] + headers

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill
            

        account_ids = report.account_ids if isinstance(report, CombinedTaxReport) else [report.account_id]

        stmt = (
            select(Gain)
            .options(joinedload(Gain.buy_lot))
            .join(Trade, Gain.sell_trade_id == Trade.id)
            .join(Account, Trade.account_id == Account.id)
            .where(Account.account_id.in_(account_ids))
            .where(Gain.tax_year == report.tax_year)
            .where(Gain.tax_pool == pool_filter)
            .order_by(Account.account_id.asc(), Trade.settle_date.asc())
        )
        gains = self.session.execute(stmt).scalars().all()
        
        for r_idx, g in enumerate(gains, 2):
            col_off = 0
            if show_account:
                ws.cell(row=r_idx, column=1).value = g.sell_trade.account.account_id
                col_off = 1

            ws.cell(row=r_idx, column=1+col_off).value = g.sell_trade.settle_date
            ws.cell(row=r_idx, column=2+col_off).value = g.buy_lot.settle_date
            ws.cell(row=r_idx, column=3+col_off).value = g.sell_trade.symbol
            
            qty_cell = ws.cell(row=r_idx, column=4+col_off)
            qty_cell.value = g.quantity_matched
            qty_cell.number_format = self.qty_format
            
            p_cell = ws.cell(row=r_idx, column=5+col_off)
            p_cell.value = g.proceeds + g.sell_comm
            p_cell.number_format = self.euro_format
            
            c_cell = ws.cell(row=r_idx, column=6+col_off)
            c_cell.value = g.cost_basis_matched - g.buy_comm
            c_cell.number_format = self.euro_format
            
            s_cell = ws.cell(row=r_idx, column=7+col_off)
            s_cell.value = g.buy_comm + g.sell_comm
            s_cell.number_format = self.euro_format

            gn_cell = ws.cell(row=r_idx, column=8+col_off)
            gn_cell.value = g.realized_pnl
            gn_cell.number_format = self.euro_format
            
        ws.freeze_panes = "A2"
        # Adjusted width for account column
        if show_account:
            ws.column_dimensions["A"].width = 15
        
        for col_l in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            real_col = chr(ord(col_l) + (1 if show_account else 0))
            if real_col <= 'I': # safety
                ws.column_dimensions[real_col].width = 15


    def _add_cash_details_sheet(self, wb, report, show_account=False):
        ws = wb.create_sheet(self.TR["sheet_dividends"])
        headers = [
            self.TR["header_date"], self.TR["header_symbol"], self.TR["header_description"], self.TR["header_type"], self.TR["header_currency"], 
            self.TR["header_amount_brutto"], self.TR["header_fx_rate"], self.TR["header_amount_eur"], self.TR["header_wht_eur"]
        ]
        if show_account:
            headers = [self.TR["header_account"]] + headers

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill
            

        # Margin/Non-taxable types to exclude from this sheet
        exclude_types = {"broker interest paid", "bond interest paid", "deposits & withdrawals"}
        account_ids = report.account_ids if isinstance(report, CombinedTaxReport) else [report.account_id]

        stmt = (
            select(CashTransaction)
            .join(Account, CashTransaction.account_id == Account.id)
            .where(Account.account_id.in_(account_ids))
            .where(CashTransaction.settle_date.like(f"{report.tax_year}%"))
            .order_by(Account.account_id.asc(), CashTransaction.settle_date.asc())
        )
        all_txs = self.session.execute(stmt).scalars().all()
        # Filter: exclude margin costs and non-taxable cash movements
        txs = [
            ct for ct in all_txs
            if ct.type.lower() not in exclude_types
            and not (ct.type.lower() == "broker interest paid/received" and ct.amount < 0)
        ]
        
        for r_idx, ct in enumerate(txs, 2):
            col_off = 0
            if show_account:
                ws.cell(row=r_idx, column=1).value = ct.account.account_id
                col_off = 1

            ws.cell(row=r_idx, column=1+col_off).value = ct.settle_date
            ws.cell(row=r_idx, column=2+col_off).value = ct.symbol or "--"
            ws.cell(row=r_idx, column=3+col_off).value = ct.description
            ws.cell(row=r_idx, column=4+col_off).value = ct.type
            ws.cell(row=r_idx, column=5+col_off).value = ct.currency
            
            amt_cell = ws.cell(row=r_idx, column=6+col_off)
            amt_cell.value = abs(ct.amount) if ct.type == "Withholding Tax" else ct.amount
            amt_cell.number_format = self.qty_format
            
            ws.cell(row=r_idx, column=7+col_off).value = ct.fx_rate_to_base
            
            eur_cell = ws.cell(row=r_idx, column=8+col_off)
            eur_amt = ct.amount * ct.fx_rate_to_base
            # If it's a dividend/interest, put it in col 8. If tax, put in col 9.
            if ct.type == "Withholding Tax":
                eur_cell.value = 0
                wht_cell = ws.cell(row=r_idx, column=9+col_off)
                wht_cell.value = abs(eur_amt)
                wht_cell.number_format = self.euro_format
            else:
                eur_cell.value = eur_amt
                eur_cell.number_format = self.euro_format
                ws.cell(row=r_idx, column=9+col_off).value = 0
            
        ws.freeze_panes = "A2"
        # Adjust column widths
        desc_col = "D" if show_account else "C"
        ws.column_dimensions[desc_col].width = 40
        
        cols_to_wide = ["A", "B", "D", "E", "F", "G", "H", "I"]
        if show_account:
            cols_to_wide.append("J")

        for col in cols_to_wide:
            if col != desc_col:
                ws.column_dimensions[col].width = 15

    def _add_margin_interest_sheet(self, wb, report, show_account=False):
        """Informational sheet for margin interest paid — not deductible per § 20 Abs. 9 EStG."""
        ws = wb.create_sheet(self.TR["sheet_margin"])

        # Header note
        ws.merge_cells("A1:G1")
        note_cell = ws["A1"]
        note_cell.value = self.TR["margin_warning"]
        note_cell.font = Font(bold=True, color="CC0000")
        note_cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[1].height = 40

        # Summary row
        ws["A3"] = self.TR["margin_total_label"]
        ws["A3"].font = self.bold_font
        summary_cell = ws["B3"]
        summary_cell.value = report.margin_interest_paid
        summary_cell.number_format = self.euro_format
        summary_cell.font = self.bold_font

        # Detail headers
        detail_headers = [
            self.TR["header_date"], self.TR["header_symbol"], self.TR["header_description"], self.TR["header_currency"],
            self.TR["header_amount_brutto"], self.TR["header_fx_rate"], self.TR["header_amount_eur"]
        ]
        if show_account:
            detail_headers = [self.TR["header_account"]] + detail_headers

        for col_idx, header in enumerate(detail_headers, 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill

        # Fetch margin interest transactions
        margin_types = {"broker interest paid", "bond interest paid"}
        account_ids = report.account_ids if isinstance(report, CombinedTaxReport) else [report.account_id]
        
        stmt = (
            select(CashTransaction)
            .join(Account, CashTransaction.account_id == Account.id)
            .where(Account.account_id.in_(account_ids))
            .where(CashTransaction.settle_date.like(f"{report.tax_year}%"))
            .order_by(Account.account_id.asc(), CashTransaction.settle_date.asc())
        )
        all_txs = self.session.execute(stmt).scalars().all()
        margin_txs = [
            ct for ct in all_txs
            if ct.type.lower() in margin_types
            or (ct.type.lower() == "broker interest paid/received" and ct.amount < 0)
        ]

        for r_idx, ct in enumerate(margin_txs, 6):
            col_off = 0
            if show_account:
                ws.cell(row=r_idx, column=1).value = ct.account.account_id
                col_off = 1

            ws.cell(row=r_idx, column=1+col_off).value = ct.settle_date
            ws.cell(row=r_idx, column=2+col_off).value = ct.symbol or "--"
            ws.cell(row=r_idx, column=3+col_off).value = ct.description
            ws.cell(row=r_idx, column=4+col_off).value = ct.currency

            amt_cell = ws.cell(row=r_idx, column=5+col_off)
            amt_cell.value = ct.amount
            amt_cell.number_format = self.qty_format

            ws.cell(row=r_idx, column=6+col_off).value = ct.fx_rate_to_base

            eur_cell = ws.cell(row=r_idx, column=7+col_off)
            eur_cell.value = abs(ct.amount * ct.fx_rate_to_base)
            eur_cell.number_format = self.euro_format

        ws.freeze_panes = "A6"
        desc_col = "D" if show_account else "C"
        ws.column_dimensions[desc_col].width = 40
        for col_l in ["A", "B", "C", "D", "E", "F", "G"]:
            real_col = col_l
            if show_account and col_l != "A":
                 # shift logic for simplicity
                 pass 
            ws.column_dimensions[col_l].width = 15
        if show_account:
             ws.column_dimensions["H"].width = 15

    def _add_deposits_withdrawals_sheet(self, wb, report, show_account=False):
        """Informational sheet for cash deposits and withdrawals."""
        ws = wb.create_sheet(self.TR["sheet_deposits"])
        headers = [self.TR["header_date"], self.TR["header_description"], self.TR["header_currency"], self.TR["header_amount"], self.TR["header_fx_rate"], self.TR["header_amount_eur"]]
        if show_account:
            headers = [self.TR["header_account"]] + headers

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill

        account_ids = report.account_ids if isinstance(report, CombinedTaxReport) else [report.account_id]

        stmt = (
            select(CashTransaction)
            .join(Account, CashTransaction.account_id == Account.id)
            .where(Account.account_id.in_(account_ids))
            .where(CashTransaction.settle_date.like(f"{report.tax_year}%"))
            .order_by(Account.account_id.asc(), CashTransaction.settle_date.asc())
        )
        all_txs = self.session.execute(stmt).scalars().all()
        dw_txs = [ct for ct in all_txs if ct.type.lower() == "deposits & withdrawals"]

        for r_idx, ct in enumerate(dw_txs, 2):
            col_off = 0
            if show_account:
                ws.cell(row=r_idx, column=1).value = ct.account.account_id
                col_off = 1

            ws.cell(row=r_idx, column=1+col_off).value = ct.settle_date
            ws.cell(row=r_idx, column=2+col_off).value = ct.description
            ws.cell(row=r_idx, column=3+col_off).value = ct.currency
            
            amt_cell = ws.cell(row=r_idx, column=4+col_off)
            amt_cell.value = ct.amount
            amt_cell.number_format = self.qty_format
            
            ws.cell(row=r_idx, column=5+col_off).value = ct.fx_rate_to_base
            
            eur_cell = ws.cell(row=r_idx, column=6+col_off)
            eur_cell.value = ct.amount * ct.fx_rate_to_base
            eur_cell.number_format = self.euro_format

        desc_col = "C" if show_account else "B"
        ws.column_dimensions[desc_col].width = 50
        for col_idx in range(1, 8 if show_account else 7):
            col_letter = chr(64 + col_idx)
            if col_letter != desc_col:
                ws.column_dimensions[col_letter].width = 15

    def _add_fx_gains_sheet(self, wb, report, show_account=False):
        ws = wb.create_sheet(self.TR["sheet_fx"])
        fx_headers = [
            self.TR["header_disposal_date"], self.TR["header_currency"], self.TR["header_acq_date"], self.TR["header_days_held"], self.TR["header_amount"],
            self.TR["header_proceeds_eur"], self.TR["header_cost_eur"], self.TR["header_gain_loss"], self.TR["header_tax_relevant"]
        ]
        if show_account:
            fx_headers = [self.TR["header_account"]] + fx_headers

        for col_idx, header in enumerate(fx_headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill
            

        account_ids = report.account_ids if isinstance(report, CombinedTaxReport) else [report.account_id]

        stmt_fx = (
            select(FXGain)
            .join(Account, FXGain.account_id == Account.id)
            .where(Account.account_id.in_(account_ids))
            .where(FXGain.disposal_date.like(f"{report.tax_year}%"))
            .order_by(Account.account_id.asc(), FXGain.disposal_date.asc())
        )
        fx_gains = self.session.execute(stmt_fx).scalars().all()
        
        for r_idx, g in enumerate(fx_gains, 2):
            col_off = 0
            if show_account:
                ws.cell(row=r_idx, column=1).value = g.account.account_id
                col_off = 1

            ws.cell(row=r_idx, column=1+col_off).value = g.disposal_date
            ws.cell(row=r_idx, column=2+col_off).value = g.fx_lot.currency
            ws.cell(row=r_idx, column=3+col_off).value = g.fx_lot.acquisition_date
            ws.cell(row=r_idx, column=4+col_off).value = g.days_held
            
            amt_cell = ws.cell(row=r_idx, column=5+col_off)
            amt_cell.value = g.amount_matched
            amt_cell.number_format = self.qty_format
            
            p_cell = ws.cell(row=r_idx, column=6+col_off)
            p_cell.value = g.disposal_proceeds_eur
            p_cell.number_format = self.euro_format
            
            c_cell = ws.cell(row=r_idx, column=7+col_off)
            c_cell.value = g.cost_basis_matched_eur
            c_cell.number_format = self.euro_format
            
            gn_cell = ws.cell(row=r_idx, column=8+col_off)
            gn_cell.value = g.realized_pnl_eur
            gn_cell.number_format = self.euro_format
            
            ws.cell(row=r_idx, column=9+col_off).value = self.TR["yes"] if g.is_taxable_section_23 else self.TR["no"]
            
        ws.freeze_panes = "A2"
        max_col_idx = 10 if show_account else 9
        for col_idx in range(1, max_col_idx + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15

    def _add_audit_trail_sheet(self, wb, report, show_account=False):
        ws = wb.create_sheet(self.TR["sheet_audit"])
        headers = [
            self.TR["header_trade_id"], self.TR["header_settle_date"], self.TR["header_symbol"], self.TR["header_cat"], self.TR["header_buy_sell"], self.TR["header_open_close"],
            self.TR["header_qty"], self.TR["header_price"], self.TR["header_currency"], self.TR["header_proceeds_eur"], self.TR["header_comm"], self.TR["header_taxes_eur"]
        ]
        if show_account:
            headers = [self.TR["header_account"]] + headers

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = self.bold_font
            cell.fill = self.header_fill
            

        account_ids = report.account_ids if isinstance(report, CombinedTaxReport) else [report.account_id]

        stmt = (
            select(Trade)
            .join(Account, Trade.account_id == Account.id)
            .where(Account.account_id.in_(account_ids))
            .where(Trade.settle_date.like(f"{report.tax_year}%"))
            .order_by(Account.account_id.asc(), Trade.settle_date.asc())
        )
        trades = self.session.execute(stmt).scalars().all()
        
        for r_idx, t in enumerate(trades, 2):
            col_off = 0
            if show_account:
                ws.cell(row=r_idx, column=1).value = t.account.account_id
                col_off = 1

            ws.cell(row=r_idx, column=1+col_off).value = t.ib_trade_id
            ws.cell(row=r_idx, column=2+col_off).value = t.settle_date
            ws.cell(row=r_idx, column=3+col_off).value = t.symbol
            ws.cell(row=r_idx, column=4+col_off).value = t.asset_category
            ws.cell(row=r_idx, column=5+col_off).value = t.buy_sell
            ws.cell(row=r_idx, column=6+col_off).value = t.open_close_indicator
            
            qty_cell = ws.cell(row=r_idx, column=7+col_off)
            qty_cell.value = t.quantity
            qty_cell.number_format = self.qty_format
            
            ws.cell(row=r_idx, column=8+col_off).value = t.trade_price
            ws.cell(row=r_idx, column=9+col_off).value = t.currency
            
            p_cell = ws.cell(row=r_idx, column=10+col_off)
            p_cell.value = t.proceeds * t.fx_rate_to_base
            p_cell.number_format = self.euro_format
            
            cm_cell = ws.cell(row=r_idx, column=11+col_off)
            cm_cell.value = t.ib_commission * t.fx_rate_to_base
            cm_cell.number_format = self.euro_format
            
            tx_cell = ws.cell(row=r_idx, column=12+col_off)
            tx_cell.value = t.taxes * t.fx_rate_to_base
            tx_cell.number_format = self.euro_format
            
        ws.freeze_panes = "A2"
        max_col_idx = 13 if show_account else 12
        for col_idx in range(1, max_col_idx + 1):
             ws.column_dimensions[chr(64 + col_idx)].width = 15

