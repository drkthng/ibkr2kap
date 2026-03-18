import os
import pytest
from decimal import Decimal
from openpyxl import load_workbook
from ibkr_tax.models.database import Account, Trade, Gain, FIFOLot, CashTransaction
from ibkr_tax.schemas.report import TaxReport
from ibkr_tax.services.excel_export import ExcelExportService

def _build_minimal_db(db_session):
    # 1. Setup Account
    acc = Account(account_id="U1234567")
    db_session.add(acc)
    db_session.commit()

    # 2. Add sample trades (Buy followed by Sell)
    # Buy AAPL
    b1 = Trade(ib_trade_id="B1", account_id=acc.id, symbol="AAPL", asset_category="STK", 
              description="Apple Inc.", trade_date="2023-12-20", settle_date="2023-12-22", 
              currency="USD", fx_rate_to_base=Decimal("0.9"), quantity=Decimal("10"), 
              trade_price=Decimal("100"), proceeds=Decimal("-1000"), buy_sell="BUY")
    
    # Buy MSFT
    b2 = Trade(ib_trade_id="B2", account_id=acc.id, symbol="MSFT", asset_category="STK", 
              description="Microsoft", trade_date="2024-01-10", settle_date="2024-01-12", 
              currency="USD", fx_rate_to_base=Decimal("0.9"), quantity=Decimal("5"), 
              trade_price=Decimal("400"), proceeds=Decimal("-2000"), buy_sell="BUY")

    # Sell AAPL
    s1 = Trade(ib_trade_id="T1", account_id=acc.id, symbol="AAPL", asset_category="STK", 
              description="Apple Inc.", trade_date="2024-01-01", settle_date="2024-01-03", 
              currency="USD", fx_rate_to_base=Decimal("0.9"), quantity=Decimal("-10"), 
              trade_price=Decimal("150"), proceeds=Decimal("1500"), buy_sell="SELL")
    
    # Sell MSFT
    s2 = Trade(ib_trade_id="T2", account_id=acc.id, symbol="MSFT", asset_category="STK", 
              description="Microsoft", trade_date="2024-06-10", settle_date="2024-06-12", 
              currency="USD", fx_rate_to_base=Decimal("0.9"), quantity=Decimal("-5"), 
              trade_price=Decimal("450"), proceeds=Decimal("2250"), buy_sell="SELL")

    db_session.add_all([b1, b2, s1, s2])
    db_session.commit()

    # 3. Add Lots
    l1 = FIFOLot(trade_id=b1.id, asset_category="STK", symbol="AAPL", settle_date="2023-12-22",
                original_quantity=Decimal("10"), remaining_quantity=Decimal("0"),
                cost_basis_total=Decimal("900"), cost_basis_per_share=Decimal("90"))
    
    l2 = FIFOLot(trade_id=b2.id, asset_category="STK", symbol="MSFT", settle_date="2024-01-12",
                original_quantity=Decimal("5"), remaining_quantity=Decimal("0"),
                cost_basis_total=Decimal("1800"), cost_basis_per_share=Decimal("360"))
    
    db_session.add_all([l1, l2])
    db_session.commit()

    # 4. Add sample Gains
    g1 = Gain(sell_trade_id=s1.id, buy_lot_id=l1.id, quantity_matched=10, tax_year=2024, 
              proceeds=Decimal("1350"), cost_basis_matched=Decimal("900"), 
              realized_pnl=Decimal("450"), tax_pool="Aktien")
    
    g2 = Gain(sell_trade_id=s2.id, buy_lot_id=l2.id, quantity_matched=5, tax_year=2024, 
              proceeds=Decimal("2025"), cost_basis_matched=Decimal("1800"), 
              realized_pnl=Decimal("225"), tax_pool="Aktien")

    db_session.add_all([g1, g2])
    db_session.commit()

    return acc

def test_export_creates_file(db_session, tmp_path):
    report = TaxReport(account_id="U1234567", tax_year=2024)
    output_file = str(tmp_path / "test_report.xlsx")
    
    service = ExcelExportService(db_session)
    service.export(report, output_file)
    
    assert os.path.exists(output_file)
    
    wb = load_workbook(output_file)
    assert "Anlage KAP Summary" in wb.sheetnames
    assert "Gains Detail" in wb.sheetnames

def test_summary_sheet_values(db_session, tmp_path):
    report = TaxReport(
        account_id="U1234567", 
        tax_year=2024,
        kap_line_7_kapitalertraege=Decimal("100.00"),
        kap_line_8_gewinne_aktien=Decimal("500.00"),
        kap_line_9_verluste_aktien=Decimal("200.00"),
        kap_line_10_termingeschaefte=Decimal("150.00"),
        kap_line_15_quellensteuer=Decimal("15.00"),
        total_realized_pnl=Decimal("600.00")
    )
    output_file = str(tmp_path / "summary_test.xlsx")
    
    service = ExcelExportService(db_session)
    service.export(report, output_file)
    
    wb = load_workbook(output_file)
    ws = wb["Anlage KAP Summary"]
    
    # Map row contents to values
    # Col A is Zeile, Col B is description, Col C is value
    data = {}
    # Iterate through a larger range to find the rows we need, stopping before empty rows or after finding specific tags
    for row in range(5, 20):
        zeile = ws.cell(row=row, column=1).value
        desc = ws.cell(row=row, column=2).value
        val = ws.cell(row=row, column=3).value
        
        if zeile:
            data[str(zeile)] = val
        elif desc and "Gesamt realisierter Gewinn/Verlust" in str(desc):
            data["TOTAL"] = val
            
    assert Decimal(str(data["7"])) == Decimal("100.00")
    assert Decimal(str(data["8"])) == Decimal("500.00")
    assert Decimal(str(data["9"])) == Decimal("200.00")
    assert Decimal(str(data["10"])) == Decimal("150.00")
    assert Decimal(str(data["15"])) == Decimal("15.00")
    assert Decimal(str(data["TOTAL"])) == Decimal("600.00")

def test_gains_detail_sheet_row_count(db_session, tmp_path):
    # Setup 2 gains in DB
    acc = _build_minimal_db(db_session)
    
    report = TaxReport(account_id=acc.account_id, tax_year=2024)
    output_file = str(tmp_path / "detail_test.xlsx")
    
    service = ExcelExportService(db_session)
    service.export(report, output_file)
    
    wb = load_workbook(output_file)
    ws = wb["Gains Detail"]
    
    # Count rows excluding header
    # max_row works if there's no trailing empty data
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    # Verify some content from the first gain (AAPL)
    # Col index 0: Verkaufsdatum, 1: Anschaffungsdatum, 2: Symbol, 3: Tax Pool
    assert rows[0][2] == "AAPL" # Symbol
    assert rows[0][3] == "Aktien" # Tax Pool
    assert rows[0][1] == "2023-12-22" # Buy Date

def test_gains_detail_sorted_by_date(db_session, tmp_path):
    # Setup data with specific dates
    acc = Account(account_id="U888")
    db_session.add(acc)
    db_session.commit()

    # T1 late in the year, T2 early
    t1 = Trade(ib_trade_id="TX1", account_id=acc.id, symbol="Z", asset_category="STK", 
              description="Z", trade_date="2024-12-01", settle_date="2024-12-03", 
              currency="EUR", fx_rate_to_base=Decimal("1"), quantity=Decimal("-1"), 
              trade_price=Decimal("100"), proceeds=Decimal("100"), buy_sell="SELL")
    
    t2 = Trade(ib_trade_id="TX2", account_id=acc.id, symbol="A", asset_category="STK", 
              description="A", trade_date="2024-01-01", settle_date="2024-01-03", 
              currency="EUR", fx_rate_to_base=Decimal("1"), quantity=Decimal("-1"), 
              trade_price=Decimal("100"), proceeds=Decimal("100"), buy_sell="SELL")

    db_session.add_all([t1, t2])
    db_session.commit()

    l1 = FIFOLot(trade_id=100, asset_category="STK", symbol="Z", settle_date="2023-01-01",
                original_quantity=Decimal("1"), remaining_quantity=Decimal("0"),
                cost_basis_total=Decimal("90"), cost_basis_per_share=Decimal("90"))
    l2 = FIFOLot(trade_id=101, asset_category="STK", symbol="A", settle_date="2023-01-01",
                original_quantity=Decimal("1"), remaining_quantity=Decimal("0"),
                cost_basis_total=Decimal("90"), cost_basis_per_share=Decimal("90"))
    db_session.add_all([l1, l2])
    db_session.commit()

    g1 = Gain(sell_trade_id=t1.id, buy_lot_id=l1.id, quantity_matched=1, tax_year=2024, 
              proceeds=Decimal("100"), cost_basis_matched=Decimal("90"), 
              realized_pnl=Decimal("10"), tax_pool="Aktien")
    
    g2 = Gain(sell_trade_id=t2.id, buy_lot_id=l2.id, quantity_matched=1, tax_year=2024, 
              proceeds=Decimal("100"), cost_basis_matched=Decimal("90"), 
              realized_pnl=Decimal("10"), tax_pool="Aktien")

    db_session.add_all([g1, g2])
    db_session.commit()

    report = TaxReport(account_id="U888", tax_year=2024)
    output_file = str(tmp_path / "sort_test.xlsx")
    
    service = ExcelExportService(db_session)
    service.export(report, output_file)
    
    wb = load_workbook(output_file)
    ws = wb["Gains Detail"]
    
    dates = [ws.cell(row=r, column=1).value for r in range(2, 4)]
    # Should be sorted: 2024-01-03 then 2024-12-03
    assert dates[0] == "2024-01-03"
    assert dates[1] == "2024-12-03"

def test_fx_gains_sheet_export(db_session, tmp_path):
    # 1. Setup Data
    acc = Account(account_id="U_FX_EXPORT_TEST")
    db_session.add(acc)
    db_session.commit()
    
    from ibkr_tax.models.database import FXFIFOLot, FXGain
    lot = FXFIFOLot(
        account_id=acc.id, currency="USD", acquisition_date="2024-01-01",
        original_amount=Decimal("1000"), remaining_amount=Decimal("0"),
        cost_basis_total_eur=Decimal("900"), cost_basis_per_unit_eur=Decimal("0.9")
    )
    db_session.add(lot)
    db_session.commit()
    
    gain = FXGain(
        account_id=acc.id, fx_lot_id=lot.id, disposal_date="2024-02-01",
        amount_matched=Decimal("1000"), disposal_proceeds_eur=Decimal("950"),
        cost_basis_matched_eur=Decimal("900"), realized_pnl_eur=Decimal("50"),
        days_held=31, is_taxable_section_23=True
    )
    db_session.add(gain)
    db_session.commit()
    
    # 2. Export
    report = TaxReport(account_id="U_FX_EXPORT_TEST", tax_year=2024)
    output_file = str(tmp_path / "fx_export_test.xlsx")
    
    service = ExcelExportService(db_session)
    service.export(report, output_file)
    
    # 3. Verify
    wb = load_workbook(output_file)
    assert "Währungsgewinne (§ 23 EStG)" in wb.sheetnames
    ws = wb["Währungsgewinne (§ 23 EStG)"]
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "2024-02-01" # Disposal Date
    assert rows[0][1] == "USD" # Currency
    assert rows[0][8] == "JA" # Taxable?

def test_gains_detail_buy_date_column(db_session, tmp_path):
    # Setup data
    acc = _build_minimal_db(db_session)
    report = TaxReport(account_id=acc.account_id, tax_year=2024)
    output_file = str(tmp_path / "buy_date_test.xlsx")
    
    service = ExcelExportService(db_session)
    service.export(report, output_file)
    
    wb = load_workbook(output_file)
    ws = wb["Gains Detail"]
    
    # Verify Headers
    assert ws.cell(row=1, column=1).value == "Verkaufsdatum"
    assert ws.cell(row=1, column=2).value == "Anschaffungsdatum"
    
    # Verify Content
    # Row 2 should be AAPL (sorted by settle_date 2024-01-03)
    # AAPL buy date from _build_minimal_db is 2023-12-22
    assert ws.cell(row=2, column=1).value == "2024-01-03"
    assert ws.cell(row=2, column=2).value == "2023-12-22"
    
    # Row 3 should be MSFT (sorted by settle_date 2024-06-12)
    # MSFT buy date from _build_minimal_db is 2024-01-12
    assert ws.cell(row=3, column=1).value == "2024-06-12"
    assert ws.cell(row=3, column=2).value == "2024-01-12"
