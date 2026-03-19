import os
import pytest
from decimal import Decimal
from openpyxl import load_workbook
from ibkr_tax.models.database import Account, Trade, Gain, FIFOLot, CashTransaction
from ibkr_tax.schemas.report import TaxReport, CombinedTaxReport
from ibkr_tax.services.excel_export import ExcelExportService

def test_export_combined_creates_file(db_session, tmp_path):
    # 1. Setup two accounts
    acc1 = Account(account_id="U111")
    acc2 = Account(account_id="U222")
    db_session.add_all([acc1, acc2])
    db_session.commit()

    # 2. Add a trade for each
    t1 = Trade(ib_trade_id="T1", account_id=acc1.id, symbol="AAPL", asset_category="STK", 
              description="Apple", trade_date="2024-01-01", settle_date="2024-01-03", 
              currency="EUR", fx_rate_to_base=Decimal("1"), quantity=Decimal("-10"), 
              trade_price=Decimal("150"), proceeds=Decimal("1500"), buy_sell="SELL")
    t2 = Trade(ib_trade_id="T2", account_id=acc2.id, symbol="MSFT", asset_category="STK", 
              description="Microsoft", trade_date="2024-01-02", settle_date="2024-01-04", 
              currency="EUR", fx_rate_to_base=Decimal("1"), quantity=Decimal("-5"), 
              trade_price=Decimal("400"), proceeds=Decimal("2000"), buy_sell="SELL")
    db_session.add_all([t1, t2])
    db_session.commit()

    # 3. Create dummy reports
    rep1 = TaxReport(account_id="U111", tax_year=2024, kap_line_8_gewinne_aktien=Decimal("100"))
    rep2 = TaxReport(account_id="U222", tax_year=2024, kap_line_8_gewinne_aktien=Decimal("200"))
    
    combined = CombinedTaxReport(
        account_ids=["U111", "U222"],
        tax_year=2024,
        kap_line_8_gewinne_aktien=Decimal("300"),
        per_account_reports=[rep1, rep2]
    )

    output_file = str(tmp_path / "combined_test.xlsx")
    service = ExcelExportService(db_session)
    
    # 4. Export
    service.export_combined(combined, output_file)
    
    assert os.path.exists(output_file)
    wb = load_workbook(output_file)
    
    # Verify Summary Sheet
    ws_sum = wb["Anlage KAP Summary"]
    assert "Kombiniert" in str(ws_sum["A1"].value)
    assert "U111, U222" in str(ws_sum["A2"].value)
    
    # Find the line 8 combined value
    # Rows start at 5
    found_combined = False
    for r in range(5, 50):
        if ws_sum.cell(row=r, column=1).value == "8":
            assert Decimal(str(ws_sum.cell(row=r, column=3).value)) == Decimal("300")
            found_combined = True
            break
    assert found_combined

    # Verify Detail Sheet has "Konto" column
    ws_gains = wb["Aktienveräußerungen (Mat.)"]
    assert ws_gains.cell(row=1, column=1).value == "Konto"
    assert ws_gains.cell(row=1, column=2).value == "Verkaufsdatum"
